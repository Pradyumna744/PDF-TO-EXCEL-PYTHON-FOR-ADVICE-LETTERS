#PDF TO EXCEL Convertion for INVOICE LETTER 
#You add palceholder according 

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pdfplumber
import openpyxl
import re
import os
from threading import Thread
import math
import time

class PDFProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("PDF to Excel Converter")
        self.root.geometry("600x400")

        # Default password
        self.pdf_password = '503376'

        # Create a frame for the animated headlines
        self.headline_frame = tk.Frame(root)
        self.headline_frame.pack(pady=10)

        # Animated Headlines
        self.headline1 = tk.Label(
            self.headline_frame,
            text="Health Care Services",
            font=("Helvetica", 16, "bold"),
            fg="#2196F3"  # Initial color
        )
        self.headline1.pack(pady=5)

        self.headline2 = tk.Label(
            self.headline_frame,
            text="Future Advice",
            font=("Helvetica", 12, "bold"),
            fg="#2196F3"  # Initial color
        )
        self.headline2.pack(pady=5)

        # Start headline animation
        self.animate_headlines()

        # Custom button style
        self.button_style = {
            'bg': '#2196F3',  # Material Design Blue
            'fg': 'white',
            'font': ('Helvetica', 10),
            'relief': 'flat',
            'padx': 15,
            'pady': 5,
            'cursor': 'hand2'
        }

        # PDF files folder location
        self.pdf_folder_label = tk.Label(root, text="PDF Files Folder Location:", font=("Helvetica", 10))
        self.pdf_folder_label.pack(pady=5)
        self.pdf_folder_entry = tk.Entry(root, width=50, font=("Helvetica", 10))
        self.pdf_folder_entry.pack(pady=5)
        self.pdf_folder_button = tk.Button(root, text="Browse", command=self.browse_pdf_folder, **self.button_style)
        self.pdf_folder_button.pack(pady=5)

        # Output Excel file storing folder location
        self.excel_folder_label = tk.Label(root, text="Output Excel File Folder Location:", font=("Helvetica", 10))
        self.excel_folder_label.pack(pady=5)
        self.excel_folder_entry = tk.Entry(root, width=50, font=("Helvetica", 10))
        self.excel_folder_entry.pack(pady=5)
        self.excel_folder_button = tk.Button(root, text="Browse", command=self.browse_excel_folder, **self.button_style)
        self.excel_folder_button.pack(pady=5)

        # Reset Password Button
        self.reset_password_button = tk.Button(
            root,
            text="Reset PDF Password",
            command=self.show_password_dialog,
            **self.button_style
        )
        self.reset_password_button.pack(pady=10)

        # Process button with special styling
        self.process_button = tk.Button(
            root,
            text="Process",
            command=self.start_processing,
            bg='#4CAF50',  # Material Design Green
            **{k: v for k, v in self.button_style.items() if k != 'bg'}
        )
        self.process_button.pack(pady=20)

        # Progress bar
        self.progress_bar = ttk.Progressbar(root, orient=tk.HORIZONTAL, length=300, mode='determinate')
        self.progress_bar.pack(pady=10)

        # Mini screen for file processing
        self.mini_screen = tk.Text(root, height=10, width=50, font=("Helvetica", 10))
        self.mini_screen.pack(pady=10)

        # Linear gradient theme
        self.style = ttk.Style()
        self.style.theme_use("clam")
        self.style.configure("TProgressbar", thickness=20, troughcolor='#f0f0f0', background='#4CAF50')

        # Bind hover effects for buttons
        for button in [self.pdf_folder_button, self.excel_folder_button, self.process_button, self.reset_password_button]:
            button.bind('<Enter>', lambda e, b=button: self.on_hover(e, b))
            button.bind('<Leave>', lambda e, b=button: self.on_leave(e, b))

    def show_password_dialog(self):
        """Show dialog for password reset"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Reset PDF Password")
        dialog.geometry("300x150")

        # Center the dialog
        dialog.transient(self.root)
        dialog.grab_set()

        # Current password
        current_pwd_label = tk.Label(dialog, text="Current Password:", font=("Helvetica", 10))
        current_pwd_label.pack(pady=5)
        current_pwd_entry = tk.Entry(dialog, show="*", font=("Helvetica", 10))
        current_pwd_entry.pack(pady=5)

        # New password
        new_pwd_label = tk.Label(dialog, text="New Password:", font=("Helvetica", 10))
        new_pwd_label.pack(pady=5)
        new_pwd_entry = tk.Entry(dialog, show="*", font=("Helvetica", 10))
        new_pwd_entry.pack(pady=5)

        def change_password():
            current = current_pwd_entry.get()
            new = new_pwd_entry.get()

            if current == self.pdf_password:
                self.pdf_password = new
                messagebox.showinfo("Success", "Password has been successfully changed!")
                dialog.destroy()
            else:
                messagebox.showerror("Error", "Current password is incorrect!")

        # Submit button
        submit_btn = tk.Button(
            dialog,
            text="Change Password",
            command=change_password,
            **self.button_style
        )
        submit_btn.pack(pady=10)

    def animate_headlines(self):
        """Animate the headlines with color and subtle movement"""
        colors = ['#2196F3', '#1976D2', '#1565C0', '#0D47A1']  # Blue color gradient
        self.color_index = 0
        self.animation_frame = 0

        def update_animation():
            # Color animation
            self.headline1.configure(fg=colors[self.color_index])
            self.headline2.configure(fg=colors[self.color_index])
            self.color_index = (self.color_index + 1) % len(colors)

            # Subtle floating animation
            y_offset = math.sin(self.animation_frame * 0.1) * 2
            self.headline1.place_configure(rely=0.5 + y_offset / 100)
            self.headline2.place_configure(rely=0.5 + y_offset / 100)

            self.animation_frame += 1
            self.root.after(100, update_animation)

        update_animation()

    def on_hover(self, event, button):
        """Button hover effect - darken the button"""
        current_bg = button.cget('bg')
        if current_bg == '#2196F3':  # Blue buttons
            button.configure(bg='#1976D2')
        elif current_bg == '#4CAF50':  # Green button
            button.configure(bg='#388E3C')

    def on_leave(self, event, button):
        """Button leave effect - restore original color"""
        current_bg = button.cget('bg')
        if current_bg == '#1976D2':  # Blue buttons
            button.configure(bg='#2196F3')
        elif current_bg == '#388E3C':  # Green button
            button.configure(bg='#4CAF50')

    def browse_pdf_folder(self):
        folder_selected = filedialog.askdirectory()
        self.pdf_folder_entry.delete(0, tk.END)
        self.pdf_folder_entry.insert(0, folder_selected)

    def browse_excel_folder(self):
        folder_selected = filedialog.askdirectory()
        self.excel_folder_entry.delete(0, tk.END)
        self.excel_folder_entry.insert(0, folder_selected)

    def start_processing(self):
        pdf_folder = self.pdf_folder_entry.get()
        excel_folder = self.excel_folder_entry.get()

        if not pdf_folder or not excel_folder:
            messagebox.showerror("Error", "Please select both PDF and Excel folder locations.")
            return

        self.process_button.config(state=tk.DISABLED)
        self.progress_bar.start()

        Thread(target=self.process_pdfs, args=(pdf_folder, excel_folder)).start()

    def process_pdfs(self, pdf_folder, excel_folder):
        pdf_files = [f for f in os.listdir(pdf_folder) if f.endswith('.pdf')]
        total_files = len(pdf_files)
        processed_files = 0

        for pdf_file in pdf_files:
            pdf_path = os.path.join(pdf_folder, pdf_file)
            self.mini_screen.insert(tk.END, f"Processing: {pdf_file}\n")
            self.mini_screen.see(tk.END)
            self.process_pdf_to_excel(pdf_path, excel_folder)
            processed_files += 1
            self.progress_bar["value"] = (processed_files / total_files) * 100
            self.root.update_idletasks()

        self.progress_bar.stop()
        self.process_button.config(state=tk.NORMAL)
        messagebox.showinfo("Completed", "PDF processing completed.")

    def process_pdf_to_excel(self, pdf_path, excel_folder):
        try:
            with pdfplumber.open(pdf_path, password=self.pdf_password) as pdf:
                text = ''
                for page in pdf.pages:
                    text += page.extract_text()
                    
            #ADD THE PLACEHOLDER IN BELOW EMPTY QUOTES 
            
            placeholders = [
                "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""
            ]

            excel_file_path = os.path.join(excel_folder, 'Invoices_Data.xlsx')

            if os.path.exists(excel_file_path):
                wb = openpyxl.load_workbook(excel_file_path)
                ws = wb.active
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(placeholders)

            data = []
            for placeholder in placeholders:
                start_index = text.find(placeholder)
                if start_index != -1:
                    start_value = text[start_index + len(placeholder):].strip()
                    next_placeholder = placeholders[placeholders.index(placeholder) + 1] if placeholders.index(placeholder) + 1 < len(placeholders) else None

                    if next_placeholder:
                        end_index = start_value.find(next_placeholder)
                        value = start_value[:end_index].strip() if end_index != -1 else start_value.strip()
                    else:
                        value = start_value.strip()

                    value = value.replace(":", "").strip()
                    data.append(value)
                else:
                    data.append("Not Found")

            data_string = " | ".join(data)

            if not os.path.exists(excel_file_path) or data_string not in [row[0] for row in ws.iter_rows(min_row=2, values_only=True)]:
                ws.append(data)
                wb.save(excel_file_path)

                self.mini_screen.insert(tk.END, f"New data added to Excel for {pdf_path}\n")
            else:
                self.mini_screen.insert(tk.END, f"Data already exists in the Excel file for {pdf_path}. No duplicate added.\n")

            self.mini_screen.see(tk.END)

        except Exception as e:
            self.mini_screen.insert(tk.END, f"Error processing PDF {pdf_path}: {str(e)}\n")
            self.mini_screen.see(tk.END)

if __name__ == "__main__":
    root = tk.Tk()
    app = PDFProcessorApp(root)
    root.mainloop()
